import tkinter as tk
from tkinter import ttk, messagebox
from dotenv import load_dotenv
import mysql.connector
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList 
from openpyxl.chart import BarChart, Reference, PieChart
load_dotenv()

senhar_db = os.getenv("DB_PASSWORD")
senhar = '3264128'
usuario = 'ma'
tentativa = 0

login_janela = tk.Tk()
login_janela.title('login')
login_janela.geometry("300x200")

root = tk.Tk()
root.title('controle de estoque')
root.geometry('600x400')

tree = ttk.Treeview(root,  columns=('id_produto', 'nome_produto', 'categoria', 'unidade','quantidade', 'preco', 'peso', 'cor'), show='headings')
tree.heading('id_produto', text='id_produto')
tree.heading('nome_produto', text='nome_produto')
tree.heading('categoria', text='categoria')
tree.heading('unidade', text='unidade')
tree.heading('quantidade', text='quantidade')
tree.heading('preco', text='preco')
tree.heading('peso', text='peso')
tree.heading('cor', text='cor')

tree.column('id_produto', width=50, anchor='center')
tree.column('nome_produto', width=200, anchor='w')
tree.column('categoria', width=200, anchor='w')
tree.column('unidade', width=200, anchor='w')
tree.column('quantidade', width=100, anchor='center')
tree.column('preco', width=50, anchor='e')
tree.column('peso', width=50, anchor='e')
tree.column('cor', width=50, anchor='w')

tree.pack(fill='both', expand=True)

tk.Label(login_janela, text='usuario:').pack(pady=5)
entrada_usuarioadm = tk.Entry(login_janela)
entrada_usuarioadm.pack()

tk.Label(login_janela, text='senha:').pack(pady=5)
entrada_sebharadm = tk.Entry(login_janela)
entrada_sebharadm.pack()

def login():
    global tentativa, conexao
    usuario_adm = entrada_usuarioadm.get()
    senhar_adm = entrada_sebharadm.get()
    
    if senhar_adm == senhar and usuario_adm == usuario:
        try:
            conexao = mysql.connector.connect(
                host="localhost",
                user="root",
                password=senhar_db,
                database="controle_estoque"
            )
            messagebox.showinfo("Acesso", "✅ Acesso autorizado")
            login_janela.destroy()
            menu()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao conectar ao banco: {e}")
    else:
        tentativa = tentativa + 1
        messagebox.showinfo("Erro", "❌ Acesso negado")
        if tentativa == 3:
            messagebox.showinfo("Bloqueado", "🚫 Conta bloqueada após 3 tentativas")
            login_janela.destroy()

tk.Button(login_janela, text='entra', command=login).pack(pady=15)

def atualizar_estoque():
    for item in tree.get_children():
        tree.delete(item)

    try:
        cursor = conexao.cursor()
        cursor.execute("SELECT id_produto, nome_produto, categoria, unidade, quantidade, preco, peso, cor FROM produto")
        produtos = cursor.fetchall()

        
        for p in produtos:
            tree.insert("", "end", values=(p[0], p[1], p[2], p[3], p[4], f'R$ {p[5]:.2f}', f'{p[6]:.2f}', p[7]))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar estoque: {e}")
btn_consuta = tk.Button(root, text='controle de estoque', command=atualizar_estoque)
btn_consuta.pack(pady=10)

def cadastrar_produto():
    janela_cadastro = tk.Toplevel()
    janela_cadastro.title('Cadastrar Produto')
    janela_cadastro.geometry('350x450')

    campos = {}
    entradas = []

    nomes = ['nome', 'categoria', 'unidade', 'quantidade', 'preco', 'peso', 'cor']

    for nome_campo in nomes:
        tk.Label(janela_cadastro, text=nome_campo.capitalize() + ':').pack()
        entrada = tk.Entry(janela_cadastro)
        entrada.pack()
        campos[nome_campo] = entrada
        entradas.append(entrada)

    
    for i in range(len(entradas) - 1):
        entradas[i].bind("<Return>", lambda event, next=entradas[i+1]: next.focus())
    entradas[-1].bind("<Return>", lambda event: salvar_produto())

    def salvar_produto():
        try:
            nome = campos['nome'].get()
            categoria = campos['categoria'].get()
            unidade = campos['unidade'].get()
            quantidade = int(campos['quantidade'].get())
            preco = float(campos['preco'].get())
            peso = float(campos['peso'].get())
            cor = campos['cor'].get()

            cursor = conexao.cursor(buffered=True)
            cursor.execute("""
                INSERT INTO produto(nome_produto, categoria, unidade, quantidade, preco, peso, cor)
                VALUES(%s, %s, %s, %s, %s, %s, %s)
            """, (nome, categoria, unidade, quantidade, preco, peso, cor))
            conexao.commit()

            messagebox.showinfo("✅ Sucesso", "Produto cadastrado com sucesso!")
            atualizar_estoque()
            janela_cadastro.destroy()

        except Exception as e:
            messagebox.showerror("❌ Erro", f"Erro ao cadastrar: {e}")

    tk.Button(janela_cadastro, text='Salvar', command=salvar_produto).pack(pady=10)
    tk.Button(janela_cadastro, text='Cancelar', command=janela_cadastro.destroy).pack()

def consutar_estoque():
    janela_consuta = tk.Toplevel()
    janela_consuta.title('consutar estoque')
    janela_consuta.geometry('600x400')

    text_area = tk.Text(janela_consuta, wrap='word')
    text_area.pack(expand=True, fill='both', padx=10, pady=10)

    try:
        cursor = conexao.cursor()
        cursor.execute('SELECT * FROM produto')
        produtos = cursor.fetchall()
        if not produtos:
            text_area.insert(tk.END, "⚠️ Nenhum produto encontrado no estoque.\n")
        else:
            for produto in produtos:
                text_area.insert(tk.END, f"🆔 ID: {produto[0]}\n")
                text_area.insert(tk.END, f"📦 Nome: {produto[1]}\n")
                text_area.insert(tk.END, f"📂 Categoria: {produto[2]}\n")
                text_area.insert(tk.END, f"📏 Unidade: {produto[3]}\n")
                text_area.insert(tk.END, f"🔢 Quantidade: {produto[4]}\n")
                text_area.insert(tk.END, f"💲 Preço: R${produto[5]:.2f}\n")
                text_area.insert(tk.END, f"⚖️ Peso: {produto[6]}\n")
                text_area.insert(tk.END, f"🎨 Cor: {produto[7]}\n")
                text_area.insert(tk.END, "-"*50 + "\n")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao consultar o estoque:\n{e}")

def alterar():
    janela_alterar = tk.Toplevel()
    janela_alterar.title("Alterar Produto")
    janela_alterar.geometry("400x500")

    tk.Label(janela_alterar, text="Digite o ID do produto:").pack(pady=5)
    entrada_id = tk.Entry(janela_alterar)
    entrada_id.pack(pady=5)

    campos = {}

    def buscar():
        try:
            id_produto = int(entrada_id.get())

            cursor = conexao.cursor()
            cursor.execute("SELECT * FROM produto WHERE id_produto = %s", (id_produto,))
            resultado = cursor.fetchone()

            if resultado:
                labels = ['nome_produto', 'categoria', 'unidade', 'quantidade', 'preco', 'peso', 'cor']
                for i, label in enumerate(labels):
                    tk.Label(janela_alterar, text=label.capitalize() + ":").pack()
                    entrada = tk.Entry(janela_alterar)
                    entrada.insert(0, str(resultado[i + 1])) 
                    entrada.pack()
                    campos[label] = entrada
            else:
                messagebox.showerror("Erro", "Produto não encontrado.")
        except ValueError:
            messagebox.showerror("Erro", "ID inválido. Digite um número inteiro.")

    def salvar():
        try:
            id_produto = int(entrada_id.get())
            dado = {
                'nome_produto': campos['nome_produto'].get(),
                'categoria': campos['categoria'].get(),
                'unidade': campos['unidade'].get(),
                'quantidade': int(campos['quantidade'].get()),
                'preco': float(campos['preco'].get()),
                'peso': float(campos['peso'].get()),
                'cor': campos['cor'].get()
            }

            cursor = conexao.cursor()
            cursor.execute("""
                UPDATE produto SET
                    nome_produto = %s,
                    categoria = %s,
                    unidade = %s,
                    quantidade = %s,
                    preco = %s,
                    peso = %s,
                    cor = %s
                WHERE id_produto = %s
            """, (
                dado['nome_produto'], dado['categoria'], dado['unidade'],
                dado['quantidade'], dado['preco'], dado['peso'], dado['cor'],
                id_produto
            ))
            conexao.commit()
            messagebox.showinfo("Sucesso", "Produto alterado com sucesso!")
            atualizar_estoque
            janela_alterar.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar alterações: {e}")

    tk.Button(janela_alterar, text="Buscar", command=buscar).pack(pady=10)
    tk.Button(janela_alterar, text="Salvar alterações", command=salvar).pack(pady=10)

def excluir():
    janela_excluir = tk.Toplevel()
    janela_excluir.title('excluir produto')
    janela_excluir.geometry('300x200')
    janela_excluir.resizable(False, False)

    tk.Label(janela_excluir, text='digte o id do produto:').pack(pady=10)
    entrada_id = tk.Entry(janela_excluir)
    entrada_id.pack(pady=5)

    def confimar():
        try:
            id_produto = int(entrada_id.get())

            cursor = conexao.cursor()

            cursor.execute('SELECT nome_produto  from produto WHERE id_produto = %s', (id_produto,))
            produto = cursor.fetchone()

            if produto:
                nome = produto[0]
                confimar = messagebox.askyesno("Confirmar", f"Tem certeza que deseja excluir o produto '{nome}'?")
                if confimar:
                    cursor.execute('DELETE FROM entrada WHERE id_produto = %s', (id_produto,))
                    cursor.execute('DELETE FROM saida WHERE id_produto = %s', (id_produto,))
                    cursor.execute('DELETE FROM produto WHERE id_produto = %s', (id_produto,))
                    cursor.execute('SELECT MAX(id_produto) from produto')
                    maior_id = cursor.fetchone()[0]
                    if maior_id is None:
                        maior_id = 0
                        cursor.execute(f'ALTER TABLE produto AUTO_INCREMENT = {maior_id + 1}')

                    conexao.commit()
                    messagebox.showinfo("Sucesso", "Produto excluído com sucesso!")
                    atualizar_estoque()
                    janela_excluir.destroy()

            else:
                messagebox.showerror("Erro", "ID de produto não encontrado.")
        except ValueError:
            messagebox.showerror("Erro", "Digite um ID válido (número inteiro).")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir produto: {e}")

    tk.Button(janela_excluir, text="Excluir", command=confimar).pack(pady=20)

def entrada():
    janela_entrada = tk.Toplevel()
    janela_entrada.title('adicionar produto')
    janela_entrada.geometry('300x200')
    janela_entrada.resizable(False, False)

    tk.Label(janela_entrada, text='digte o id do produto:').pack(pady=10)
    entrada_id = tk.Entry(janela_entrada)
    entrada_id.pack(pady=5)

    tk.Label(janela_entrada, text='digite a quantidade de entrada').pack(pady=10)
    entrada_quantidade = tk.Entry(janela_entrada)
    entrada_quantidade.pack(pady=5)

    def confirmar():
        try:
            id_produto = int(entrada_id.get())
            quantidade = int(entrada_quantidade.get())

            cursor = conexao.cursor()

            cursor.execute("SELECT nome_produto FROM produto WHERE id_produto = %s", (id_produto,))
            produto = cursor.fetchone()

            if not produto:
                messagebox.showerror("Erro", "Produto não encontrado.")
                return

            nome = produto[0]

            cursor.execute('''
        INSERT INTO entrada(id_produto, quantidade_entrada, data_entrada)
        VALUES(%s,%s, NOW())
        ''', (id_produto, quantidade))
            
            cursor.execute('''
        UPDATE produto
        SET quantidade = quantidade + %s 
        WHERE id_produto = %s 
        ''', (quantidade, id_produto))
            
            conexao.commit()
            messagebox.showinfo("Sucesso", f"Entrada de {quantidade} unidade(s) do produto '{nome}' registrada!")
            atualizar_estoque()
            janela_entrada.destroy()
        except ValueError:
            messagebox.showerror("Erro", "Digite valores numéricos válidos.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao registrar entrada: {e}")
           
    tk.Button(janela_entrada, text="Confirmar", command=confirmar).pack(pady=20)
def saida():
    janela_saida = tk.Toplevel()
    janela_saida.title('retirada de produto')
    janela_saida.geometry('300x200')
    janela_saida.resizable(False, False)

    tk.Label(janela_saida, text='digte o id do produto:').pack(pady=10)
    entrada_id = tk.Entry(janela_saida)
    entrada_id.pack(pady=5)

    tk.Label(janela_saida, text='digite a quantidade de entrada').pack(pady=10)
    entrada_quantidade = tk.Entry(janela_saida)
    entrada_quantidade.pack(pady=5)

    def confirmar():
        try:
            id_produto = int(entrada_id.get())
            quantidade = int(entrada_quantidade.get())

            cursor = conexao.cursor()

            cursor.execute("SELECT nome_produto FROM produto WHERE id_produto = %s", (id_produto,))
            produto = cursor.fetchone()

            if not produto:
                messagebox.showerror("Erro", "Produto não encontrado.")
                return

            nome = produto[0]

            cursor.execute('''
        INSERT INTO saida(id_produto, quantidade_saida, data_saida)
        VALUES(%s,%s, NOW())
        ''', (id_produto, quantidade))
            
            cursor.execute('''
        UPDATE produto
        SET quantidade = quantidade - %s 
        WHERE id_produto = %s 
        ''', (quantidade, id_produto))
            
            conexao.commit()
            messagebox.showinfo("Sucesso", f"saida de {quantidade} unidade(s) do produto '{nome}' registrada!")
            atualizar_estoque()
            janela_saida.destroy()
        except ValueError:
            messagebox.showerror("Erro", "Digite valores numéricos válidos.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao registrar entrada: {e}")
           
    tk.Button(janela_saida, text="Confirmar", command=confirmar).pack(pady=20)

def gerar_relatorio():
    try:
        cursor = conexao.cursor()

        cursor.execute('''
            SELECT p.id_produto,p.nome_produto,
            IFNULL(SUM(e.quantidade_entrada),0) - IFNULL(SUM(s.quantidade_saida),0) AS estoque_atual
            FROM produto p
            LEFT JOIN entrada e on p.id_produto = e.id_produto
            LEFT JOIN saida s on p.id_produto = s.id_produto
            GROUP BY p.id_produto, p.nome_produto
                       ''')
        estoque = cursor.fetchall()

        cursor.execute('''
        SELECT e.id_entrada, p.nome_Produto, e.quantidade_entrada, e.data_entrada
        FROM entrada e
        INNER JOIN produto p ON e.id_produto = p.id_produto
        ORDER BY e.data_entrada DESC
        ''')
        entradas = cursor.fetchall()

        cursor.execute('''
        SELECT s.id_saida, p.nome_produto, s.quantidade_saida, s.data_saida
        FROM saida s
        INNER JOIN produto p on s.id_produto = p.id_produto
        ORDER BY s.data_saida DESC
        ''')
        saidas = cursor.fetchall()

        cursor.execute('''
        SELECT id_despesa, descricao, valor, data_despesa
        FROM despesa
        ORDER BY data_despesa DESC
        ''')
        despesas = cursor.fetchall()

        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = 'estoque'
        ws1.append(['id_produto', 'produto', 'quantidade'])
        for linha in estoque:
            ws1.append(linha)
        if len(estoque) > 0:
            chart1 = BarChart()
            chart1.title = 'estoque'
            dados1 = Reference(ws1, min_col=3, min_row =1, max_row=len(estoque) + 1)
            categoria1 = Reference(ws1, min_col=2, min_row=2, max_row=len(estoque) + 1)
            chart1.add_data(dados1, titles_from_data=True)
            chart1.set_categories(categoria1)
            chart1.y_axis.title = 'quantidade'
            chart1.x_axis.title = 'produto'
            chart1.dLbls = DataLabelList()
            chart1.dLbls.showVal = True
            chart1.series[0].graphicalProperties.solidfill = '1F77B4'
            ws1.add_chart(chart1, 'E2')
            wb1.save('estoque.xlsx')

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = 'entrada'
        ws2.append(['id_entrada', 'produto', 'quantidade_entrada', 'data_entrada'])
        for linha in entradas:
            ws2.append(linha)
        if len(entradas) > 0:
            chart2 = BarChart()
            chart2.title = 'entrada'
            dados2 = Reference(ws2, min_col=3, min_row =1, max_row=len(entradas) + 1)
            categoria2 = Reference(ws2, min_col=2, min_row=2, max_row=len(entradas) + 1)
            chart2.add_data(dados2, titles_from_data=True)
            chart2.set_categories(categoria2)
            chart2.y_axis.title = 'quantidade_entrada'
            chart2.x_axis.title = 'produto'
            chart2.dLbls = DataLabelList()
            chart2.dLbls.showVal = True
            chart2.series[0].graphicalProperties.solidfill = '2CA02C'
            ws2.add_chart(chart2, 'F2')
            wb2.save('entrada.xlsx')

        wb3 = Workbook()
        ws3 = wb3.active
        ws3.title = 'saida'
        ws3.append(['id_saida', 'produto', 'quantidade_saida', 'data_saida'])
        for linha in saidas:
            ws3.append(linha)
        if len(saidas) > 0:
            chart3 = BarChart()
            chart3.title = 'saida'
            dados3 = Reference(ws3, min_col=3, min_row =1, max_row=len(saidas) + 1)
            categoria3 = Reference(ws3, min_col=2, min_row=2, max_row=len(saidas) + 1)
            chart3.add_data(dados3, titles_from_data=True)
            chart3.set_categories(categoria3)
            chart3.y_axis.title = 'quantidade_saida'
            chart3.x_axis.title = 'produto'
            chart3.dLbls = DataLabelList()
            chart3.dLbls.showVal = True
            chart3.series[0].graphicalProperties.solidfill = 'D62728'
            ws3.add_chart(chart3, 'F2')
            wb3.save('saida.xlsx')

        wb4 = Workbook()
        ws4 = wb4.active
        ws4.title = 'despesa'
        ws4.append(['id_despesa', 'descricao', 'valor', 'data_despesa'])
        for linha in despesas:
            ws4.append(linha)
        if len(despesas) > 0:
            chart4 = BarChart()
            chart4.title = 'despesa'
            dados4 = Reference(ws4, min_col=3, min_row =1, max_row=len(despesas) + 1)
            categoria4 = Reference(ws4, min_col=2, min_row=2, max_row=len(despesas) + 1)
            chart4.add_data(dados4, titles_from_data=True)
            chart4.set_categories(categoria4)
            chart4.y_axis.title = 'valor R$'
            chart4.x_axis.title = 'descricao'
            chart4.dLbls = DataLabelList()
            chart4.dLbls.showVal = True
            chart4.series[0].graphicalProperties.solidfill = 'FF7F0E'
            ws4.add_chart(chart4, 'F2')
            wb4.save('despesa.xlsx')
            

        messagebox.showinfo("Relatório", f"Relatório completo gerado!\nestoque.xlsx, entrada.xlsx, saida.xlsx, despesa.xlsx")
        cursor.close()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar relatório completo: {e}")

def gerar_relatorio_periodo():
    janela_data = tk.Toplevel()
    janela_data.title("Selecionar Período")
    janela_data.geometry("300x150")

    tk.Label(janela_data, text="Data Inicial yyyy-mm-dd:").pack(pady=5)
    entry_inicio = tk.Entry(janela_data)
    entry_inicio.pack()

    tk.Label(janela_data, text="Data Final yyyy-mm-dd:").pack(pady=5)
    entry_fim = tk.Entry(janela_data)
    entry_fim.pack()

    

    def confirmar():
        try:
            def parse_data(valor):
                valor = valor.strip()
                for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                    try:
                        return datetime.strptime(valor, fmt).date()
                    except ValueError:
                        pass
                raise ValueError("Formato de data inválido")

            data_inicial = parse_data(entry_inicio.get())
            data_final = parse_data(entry_fim.get())
        

            if data_inicial > data_final:
                messagebox.showwarning("Aviso", "A data inicial não pode ser maior que a final.")
                return
            cursor = conexao.cursor()
            cursor.execute('''
            SELECT p.id_produto, p.nome_produto,
                SUM(IFNULL(e.quantidade_entrada, 0)) AS total_entrada,
                SUM(IFNULL(s.quantidade_saida, 0)) AS total_saida
                FROM produto p
                LEFT JOIN entrada e  ON p.id_produto = e.id_produto AND e.data_entrada BETWEEN %s AND %s
                LEFT JOIN saida s  ON p.id_produto = s.id_produto AND s.data_saida BETWEEN %s AND %s
                GROUP BY p.id_produto, p.nome_produto
                 ''', (data_inicial, data_final, data_inicial, data_final))
            resultado = cursor.fetchall()

            if resultado:
                relatorio = f"📊 Relatório de {data_inicial} até {data_final}\n"
                relatorio += "ID | Produto | Entradas | Saídas\n"
                relatorio += "-"*50 + "\n"
                for row in resultado:
                    relatorio += f"{row[0]} | {row[1]} | {row[2]} | {row[3]}\n"

                messagebox.showinfo("Relatório por Período", relatorio)
            else:
                messagebox.showinfo("Relatório por Período", "Nenhum dado encontrado no período.")

        except ValueError:
            messagebox.showerror("Erro", "Formato de data inválido! Use YYYY-MM-DD.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório por período:\n{e}")

    tk.Button(janela_data, text="Gerar", command=confirmar).pack(pady=10)

def despesa():
    janela_despesa = tk.Toplevel()
    janela_despesa.title('despesa')
    janela_despesa.geometry('300x150')

    tk.Label(janela_despesa, text='descriçao:').pack(pady=5)
    entry_descricao = tk.Entry(janela_despesa, width=30)
    entry_descricao.pack(pady=5)

    tk.Label(janela_despesa, text='valor R$:').pack(pady=5)
    entry_valor = tk.Entry(janela_despesa, width=30)
    entry_valor.pack(pady=5)

    tk.Label(janela_despesa, text='data YYYY-MM-DD').pack(pady=5)
    entry_data = tk.Entry(janela_despesa, width=30)
    entry_data.pack(pady=5)

    def salvar():
        try:
            descriçao = entry_descricao.get()
            valor = entry_valor.get()
            data_despesa = entry_data.get()

            cursor = conexao.cursor()
            cursor.execute('''
            INSERT INTO despesa(descricao, valor, data_despesa)
            VALUES(%s, %s, %s)
                           
                           ''', (descriçao, valor, data_despesa))
            conexao.commit()
            messagebox.showinfo("Sucesso", "Despesa cadastrada com sucesso!")
            janela_despesa.destroy()
        except ValueError:
            messagebox.showerror("Erro", "Digite um valor numérico válido.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao cadastrar despesa:\n{e}")

    tk.Button(janela_despesa, text="Salvar", command=salvar).pack(pady=10)

def lista_despesa():
    janela_lista = tk.Toplevel()
    janela_lista.title('lista despesa')
    janela_lista.geometry('300x150')

    colunas = ('id_despesa', 'descricao', 'valor', 'data_despesa')
    tree = ttk.Treeview(janela_lista, columns=colunas, show="headings")
    tree.pack(fill="both", expand=True)

    tree.heading("id_descricao", text="id_despesa")
    tree.heading("descrição", text="descrição")
    tree.heading("valor", text="valor (R$)")
    tree.heading("data_despesa", text="data_despesa")

    tree.column("id_despesa", width=50)
    tree.column("descrição", width=200)
    tree.column("valor", width=100)
    tree.column("data_despesa", width=100)

    try:
        cursor = conexao.cursor()
        cursor.execute('SELECT id_despesa, descricao, valor, data_despesa FROM despesa ORDER BY data_despesa DESC')
        registro = cursor.fetchall

        for linha in registro:
            tree.insert("", "end", values=linha)

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar despesas:\n{e}")

    
    tk.Button(janela_lista, text="Fechar", command=janela_lista.destroy).pack(pady=5)


def menu():
    menu = tk.Tk()
    menu.title(menu)
    menu.geometry("300x200")
    tk.Label(menu, text='controle de estoque').pack(pady=20)
    tk.Button(menu, text='cadastar produto', command=cadastrar_produto).pack(pady=5)
    tk.Button(menu, text='consutar estoque', command=consutar_estoque).pack(pady=5)
    tk.Button(menu, text='alterar produto', command=alterar).pack(pady=5)
    tk.Button(menu, text='excluir produto', command=excluir).pack(pady=5)
    tk.Button(menu, text='adiconar produto', command=entrada).pack(pady=5)
    tk.Button(menu, text= 'retira produto', command=saida).pack(pady=5)
    tk.Button(menu, text= 'gerar relatorio', command=gerar_relatorio).pack(pady=5)
    tk.Button(menu, text='gerar relatorio por periodo', command=gerar_relatorio_periodo).pack(pady=5)
    tk.Button(menu, text='cadastra despesa', command=despesa).pack(pady=5)
    tk.Button(menu, text='lista despesa', command=lista_despesa).pack(pady=5)
    tk.Button(menu, text='sair', command=menu.destroy).pack(pady=10)
    menu.mainloop()

login_janela.mainloop()

